import pandas as pd
import time
import re
import os
import PySimpleGUI as sg
from urllib.request import urlretrieve
from openpyxl import load_workbook
from selenium import webdriver as wb
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from openpyxl.worksheet.properties import WorksheetProperties as wp
from selenium.webdriver.common.action_chains import ActionChains
from urllib.error import HTTPError


def Build_tool(sku_list, id_2, id_1):

    # Setting up the webdriver for Selenium
    options = wb.ChromeOptions()
    options.add_argument('--start-maximized')
    options.add_experimental_option('excludeSwitches', ['enable-logging'])
    driver = wb.Chrome(options=options)

    data_dict = {'Sku': [],
                 'Category': [],  
                'Img_url1': [], 
                'Img_url2': [], 
                'Img_url3': [], 
                'Img_url4': [], 
                'Img_url5': [], 
                'Img_url6': [], 
                'Img_url7': [], 
                'Img_url8': [], 
                'Img_url9': [], 
                'Bullet1': [], 
                'Bullet2': [], 
                'Bullet3': [], 
                'Bullet4': [], 
                'Bullet5': [], 
                'Bullet6' : [], 
                'Bullet7': [], 
                'Bullet8': [], 
                'Bullet9': [], 
                'PDF_1': [],
                'PDF_2': [],
                'PDF_3': [],
                'PDF_4': [], 
                'Skus_Not_Found': []}

    for (id_2, id_1, sku) in zip(id_2, id_1, sku_list):
        try:
            path = 'https://www.build.com/pfister-hhl-089tb/s{}?uid={}'.format(id_1, id_2)
            driver.get(path)
            WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "root")))
            # time.sleep(1)
            driver.execute_script("window.scrollTo(0, 300)")
            
            # Adding sku to dictionary
            data_dict['Sku'].append(sku)

            # Extracting the category
            try:
                category = driver.find_element_by_xpath("//*[@id='main-content']/div/section[1]/div[1]/nav/ol/li[4]/a/span").text
                data_dict['Category'].append(category)
            except:
                data_dict['Category'].append('NULL')

            # Extracting the image src
            src_1 = driver.find_element_by_xpath("//*[contains(@class,'w-auto self-center undefined')]").get_attribute('src')
            # time.sleep(1)
            data_dict['Img_url1'].append(src_1)
            
            try:
                element = driver.find_element_by_xpath("//div[@aria-label='thumb slide 1']")
                actions = ActionChains(driver)
                actions.move_to_element(element).perform()
                element.click()
                time.sleep(1)
                src_2 = driver.find_element_by_xpath("//div[starts-with(@aria-label, '2 /')]").find_element_by_tag_name('img').get_attribute('src')
                data_dict['Img_url2'].append(src_2)
            except:
                data_dict['Img_url2'].append('NULL')
                
            try:    
                element = driver.find_element_by_xpath("//div[@aria-label='thumb slide 2']")
                actions = ActionChains(driver)
                actions.move_to_element(element).perform()
                element.click()
                time.sleep(1)
                src_3 = driver.find_element_by_xpath("//div[starts-with(@aria-label, '3 /')]").find_element_by_tag_name('img').get_attribute('src')
                data_dict['Img_url3'].append(src_3)
            except:
                data_dict['Img_url3'].append('NULL')

            try:   
                element = driver.find_element_by_xpath("//div[@aria-label='thumb slide 3']")
                actions = ActionChains(driver)
                actions.move_to_element(element).perform()
                element.click()
                time.sleep(1)
                src_4 = driver.find_element_by_xpath("//div[starts-with(@aria-label, '4 /')]").find_element_by_tag_name('img').get_attribute('src')
                data_dict['Img_url4'].append(src_4)
            except:
                data_dict['Img_url4'].append('NULL')

            try:   
                element = driver.find_element_by_xpath("//div[@aria-label='thumb slide 4']")
                actions = ActionChains(driver)
                actions.move_to_element(element).perform()
                element.click()
                time.sleep(1)
                src_5 = driver.find_element_by_xpath("//div[starts-with(@aria-label, '5 /')]").find_element_by_tag_name('img').get_attribute('src')
                data_dict['Img_url5'].append(src_5)
            except:
                data_dict['Img_url5'].append('NULL')

            try:   
                element = driver.find_element_by_xpath("//div[@aria-label='thumb slide 5']")
                actions = ActionChains(driver)
                actions.move_to_element(element).perform()
                element.click()
                time.sleep(1)
                src_6 = driver.find_element_by_xpath("//div[starts-with(@aria-label, '6 /')]").find_element_by_tag_name('img').get_attribute('src')
                data_dict['Img_url6'].append(src_6)
            except:
                data_dict['Img_url6'].append('NULL')
            
            try:   
                element = driver.find_element_by_xpath("//div[@aria-label='thumb slide 6']")
                actions = ActionChains(driver)
                actions.move_to_element(element).perform()
                element.click()
                time.sleep(1)
                src_7 = driver.find_element_by_xpath("//div[starts-with(@aria-label, '7 /')]").find_element_by_tag_name('img').get_attribute('src')
                data_dict['Img_url7'].append(src_7)
            except:
                data_dict['Img_url7'].append('NULL')
            
            try:   
                element = driver.find_element_by_xpath("//div[@aria-label='thumb slide 7']")
                actions = ActionChains(driver)
                actions.move_to_element(element).perform()
                element.click()
                time.sleep(1)
                src_8 = driver.find_element_by_xpath("//div[starts-with(@aria-label, '8 /')]").find_element_by_tag_name('img').get_attribute('src')
                data_dict['Img_url8'].append(src_8)
            except:
                data_dict['Img_url8'].append('NULL')
            
            try:   
                element = driver.find_element_by_xpath("//div[@aria-label='thumb slide 8']")
                actions = ActionChains(driver)
                actions.move_to_element(element).perform()
                element.click()
                time.sleep(1)
                src_9 = driver.find_element_by_xpath("//div[starts-with(@aria-label, '9 /')]").find_element_by_tag_name('img').get_attribute('src')
                data_dict['Img_url9'].append(src_9)
            except:
                data_dict['Img_url9'].append('NULL')

            # Extracting bullet points
            try:
                element = driver.find_element_by_xpath("//div[@class='lh-copy H_oFW']")
                # element = driver.find_element_by_xpath("//*[@id='main-content']/div/section[4]/div[1]/section/div[1]/section/div[3]/section/div/div[1]/div[1]/ul[1]")
                actions = ActionChains(driver)
                actions.move_to_element(element).perform()
                # driver.execute_script("window.scrollTo(0, 800)")
                time.sleep(3)
                
                # Extracting the webelement then transforming it to text with no special characters
                li_elements = element.find_elements_by_tag_name('li')
                b_list = []
                
                for li in li_elements:
                    li_text = li.text
                    bullet = re.sub("[^A-Za-z0-9 -\/]", "", li_text)
                    bullet = bullet.replace('"', "-in") 
                    b_list.append(bullet)
                
                # Loading all bullet points found into the data_dict
                try:
                    data_dict['Bullet1'].append(b_list[0])
                except:
                    data_dict['Bullet1'].append('NULL')
                
                try:
                    data_dict['Bullet2'].append(b_list[1])
                except:
                    data_dict['Bullet2'].append('NULL')
                
                try:
                    data_dict['Bullet3'].append(b_list[2])
                except:
                    data_dict['Bullet3'].append('NULL')
                
                try:
                    data_dict['Bullet4'].append(b_list[3])
                except:
                    data_dict['Bullet4'].append('NULL')

                try:
                    data_dict['Bullet5'].append(b_list[4])
                except:
                    data_dict['Bullet5'].append('NULL')
                
                try:
                    data_dict['Bullet6'].append(b_list[5])
                except:
                    data_dict['Bullet6'].append('NULL')
                
                try:
                    data_dict['Bullet7'].append(b_list[6])
                except:
                    data_dict['Bullet7'].append('NULL')
                
                try:
                    data_dict['Bullet8'].append(b_list[7])
                except:
                    data_dict['Bullet8'].append('NULL')
                
                try:
                    data_dict['Bullet9'].append(b_list[8])
                except:
                    data_dict['Bullet9'].append('NULL')
                    
            # If sku is not found
            except:
                # data_dict['Sku'].append(sku)
                data_dict['Bullet1'].append('NULL')
                data_dict['Bullet2'].append('NULL')
                data_dict['Bullet3'].append('NULL')
                data_dict['Bullet4'].append('NULL')
                data_dict['Bullet5'].append('NULL')
                data_dict['Bullet6'].append('NULL')
                data_dict['Bullet7'].append('NULL')
                data_dict['Bullet8'].append('NULL')
                data_dict['Bullet9'].append('NULL')
                # data_dict['Skus_Not_Found'].append(sku)

            # Extracting pdfs
            driver.execute_script("window.scrollTo(0, 600)")
            time.sleep(1)

            hrefs = driver.find_elements_by_xpath("//a[@class='f-inherit fw-inherit link theme-primary  pb3 f7 db underline-hover']")
            href_list = []

            for href in hrefs:
                href = href.get_attribute('href')
                href_list.append(href)
                
            try:
                data_dict['PDF_1'].append(href_list[0])
            except:
                data_dict['PDF_1'].append('NULL')
            
            try:
                data_dict['PDF_2'].append(href_list[1])
            except:
                data_dict['PDF_2'].append('NULL')

            try:
                data_dict['PDF_3'].append(href_list[2])
            except:
                data_dict['PDF_3'].append('NULL')

            try:
                data_dict['PDF_4'].append(href_list[3])
            except:
                data_dict['PDF_4'].append('NULL')

        except:
            # data_dict['Sku'].append(sku)       
            data_dict['Skus_Not_Found'].append(sku)
            data_dict['Category'].append('NULL')
            data_dict['Img_url1'].append('NULL')
            data_dict['Img_url2'].append('NULL')
            data_dict['Img_url3'].append('NULL')
            data_dict['Img_url4'].append('NULL')
            data_dict['Img_url5'].append('NULL')
            data_dict['Img_url6'].append('NULL')
            data_dict['Img_url7'].append('NULL')
            data_dict['Img_url8'].append('NULL')
            data_dict['Img_url9'].append('NULL')
            data_dict['Bullet1'].append('NULL')
            data_dict['Bullet3'].append('NULL')
            data_dict['Bullet4'].append('NULL')
            data_dict['Bullet5'].append('NULL')
            data_dict['Bullet6'].append('NULL')
            data_dict['Bullet7'].append('NULL')
            data_dict['Bullet2'].append('NULL')
            data_dict['Bullet8'].append('NULL')
            data_dict['Bullet9'].append('NULL')
            data_dict['PDF_1'].append('NULL')
            data_dict['PDF_2'].append('NULL')
            data_dict['PDF_3'].append('NULL')
            data_dict['PDF_4'].append('NULL')

    # quitting the driver and manipulation the dictionary into a dataframe
    driver.quit()

    df = pd.DataFrame.from_dict(data_dict,orient='index')
    df = df.transpose()
    # # df['Img_url'].fillna('NULL', inplace=True)

    # Writing the dataframe to an excel worksheet
    # path = r'C:\Users\ABG2137\OneDrive - MyDigitalSpace\Desktop\Image_Data.xlsx'
    # excel_wb = load_workbook(path)
    df.to_excel('Build_Data.xlsx', sheet_name='Build_Data')

    # sg.popup("Run Complete!")

def Ferg_Tool(alt_sku_list):

    # Setting up the webdriver for Selenium
    options = wb.ChromeOptions()
    options.add_argument('--start-maximized')
    options.add_experimental_option('excludeSwitches', ['enable-logging'])
    driver = wb.Chrome(options=options)

    data_dict = {'Alt_Sku': [], 
                 'Img_url1': [], 
                 'Product_Title': [], 
                 'Category': [], 
                 'PDF_1': [], 
                 'PDF_2': [], 
                 'PDF_3': [], 
                 'PDF_4': [], 
                 'Bullet1': [], 
                 'Bullet2': [], 
                 'Bullet3': [], 
                 'Bullet4': [], 
                 'Bullet5': [], 
                 'Bullet6': [], 
                 'Bullet7': [], 
                 'Bullet8': [], 
                 'Bullet9': [], 
                 'Alt_Sku_Not_Found': []}

    for alt_sku in alt_sku_list:
        try:
            path = 'https://www.ferguson.com/'
            driver.get(path)
            time.sleep(1)
            WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.NAME, "search")))
            time.sleep(1)
            driver.find_element(By.NAME, "search").click()
            time.sleep(1)
            driver.find_element(By.NAME, "search").send_keys(alt_sku)
            driver.find_element(By.NAME, "search").send_keys(Keys.RETURN)
            # WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//*[contains(@class,'hero__img')]")))
            time.sleep(5)

            # Adding alt 1 to dictionary
            data_dict['Alt_Sku'].append(alt_sku)

            # Extracting the image src
            try:
                src_1 = driver.find_element_by_xpath("//img[@class='lazyload js-img-item']").get_attribute('src')
                data_dict['Img_url1'].append(src_1)
            except:
                data_dict['Img_url1'].append('NULL')

            # Extracting the product title
            try:
                title = driver.find_element_by_xpath("//*[contains(@class,'product__name')]").text
                data_dict['Product_Title'].append(title)
            except:
                data_dict['Product_Title'].append('NULL')

            # Extracting the category
            try:
                category = driver.find_element_by_xpath("//*[@id='wrapper']/main/div/div/div[1]/div[1]/div/li[4]/a/span").text
                data_dict['Category'].append(category)
            except:
                data_dict['Category'].append('NULL')

            # Extracting pdfs
            driver.execute_script("window.scrollTo(0, 200)")
            time.sleep(1)

            hrefs = driver.find_elements_by_xpath("//a[@class='fc__blue js-doc-download']")
            href_list = []

            for href in hrefs:
                href = href.get_attribute('href')
                href_list.append(href)
                
            try:
                data_dict['PDF_1'].append(href_list[0])
            except:
                data_dict['PDF_1'].append('NULL')
            
            try:
                data_dict['PDF_2'].append(href_list[1])
            except:
                data_dict['PDF_2'].append('NULL')

            try:
                data_dict['PDF_3'].append(href_list[2])
            except:
                data_dict['PDF_3'].append('NULL')

            try:
                data_dict['PDF_4'].append(href_list[3])
            except:
                data_dict['PDF_4'].append('NULL')

            # Extracting bullet points
            try:
                element = driver.find_element_by_xpath("//*[contains(@class,'product__details__content')]")
                li_elements = element.find_elements_by_tag_name('li')
                b_list = []
                
                for li in li_elements:
                    li_text = li.text
                    bullet = re.sub("[^A-Za-z0-9 -\/]", "", li_text)
                    bullet = bullet.replace('"', "-in") 
                    b_list.append(bullet)

                # Loading all bullet points found into the data_dict
                try:
                    data_dict['Bullet1'].append(b_list[0])
                except:
                    data_dict['Bullet1'].append('NULL')
                
                try:
                    data_dict['Bullet2'].append(b_list[1])
                except:
                    data_dict['Bullet2'].append('NULL')
                
                try:
                    data_dict['Bullet3'].append(b_list[2])
                except:
                    data_dict['Bullet3'].append('NULL')
                
                try:
                    data_dict['Bullet4'].append(b_list[3])
                except:
                    data_dict['Bullet4'].append('NULL')

                try:
                    data_dict['Bullet5'].append(b_list[4])
                except:
                    data_dict['Bullet5'].append('NULL')
                
                try:
                    data_dict['Bullet6'].append(b_list[5])
                except:
                    data_dict['Bullet6'].append('NULL')
                
                try:
                    data_dict['Bullet7'].append(b_list[6])
                except:
                    data_dict['Bullet7'].append('NULL')
                
                try:
                    data_dict['Bullet8'].append(b_list[7])
                except:
                    data_dict['Bullet8'].append('NULL')
                
                try:
                    data_dict['Bullet9'].append(b_list[8])
                except:
                    data_dict['Bullet9'].append('NULL')
            except:
                data_dict['Bullet1'].append('NULL')
                data_dict['Bullet2'].append('NULL')
                data_dict['Bullet3'].append('NULL')
                data_dict['Bullet4'].append('NULL')
                data_dict['Bullet5'].append('NULL')
                data_dict['Bullet6'].append('NULL')
                data_dict['Bullet7'].append('NULL')
                data_dict['Bullet8'].append('NULL')
                data_dict['Bullet9'].append('NULL')

        except:
            driver.find_element(By.NAME, "search").clear()
            data_dict['Img_url1'].append('NULL')
            data_dict['Alt_Sku'].append(alt_sku)
            data_dict['Product_Title'].append('NULL')
            data_dict['Category'].append('NULL')
            data_dict['PDF_1'].append('NULL')
            data_dict['PDF_2'].append('NULL')
            data_dict['PDF_3'].append('NULL')
            data_dict['PDF_4'].append('NULL')
            data_dict['Bullet1'].append('NULL')
            data_dict['Bullet2'].append('NULL')
            data_dict['Bullet3'].append('NULL')
            data_dict['Bullet4'].append('NULL')
            data_dict['Bullet5'].append('NULL')
            data_dict['Bullet6'].append('NULL')
            data_dict['Bullet7'].append('NULL')
            data_dict['Bullet8'].append('NULL')
            data_dict['Bullet9'].append('NULL')
            data_dict['Alt_Sku_Not_Found'].append(alt_sku)

        
    driver.quit()

    df = pd.DataFrame.from_dict(data_dict,orient='index')
    df = df.transpose()
    # # df['Img_url'].fillna('NULL', inplace=True)

    # Writing the dataframe to an excel worksheet
    # path = r'C:\Users\ABG2137\OneDrive - MyDigitalSpace\Desktop\Image_Data.xlsx'
    # excel_wb = load_workbook(path)
    df.to_excel('Ferg_Site_Data.xlsx', sheet_name='Ferg_Site_Data')

    # sg.popup("Run Complete!")

def converter_tool(mfg_list_primary, Primary_list, folder_name):
    # Create a dictionary from the two lists for a loop
    primary_img_dict = {mfg_list_primary[i]: Primary_list[i] for i in range(len(Primary_list))}
    # img_2_dict = {mfg_list_2[i]: img_2_list[i] for i in range(len(img_2_list))}
    # img_3_dict = {mfg_list_3[i]: img_3_list[i] for i in range(len(img_3_list))}
    # img_4_dict = {mfg_list_4[i]: img_4_list[i] for i in range(len(img_4_list))}


    # Creating a folder variable for output
    output_directory = '{}'.format(folder_name)

    # Looping through the dictionary and creating .jpgs from the urls and loading the file names into a list
    # file_name_dict = {'Primary_File_name':[], 'Image_2_Name': [], 'Image_3_Name': [], 'Image_4_Name': [], '404_error_images': []}
    file_name_dict = {'Primary_File_name':[], '404_error_images': []}

    for mfg, url in primary_img_dict.items():
        try:
            primary_file_name = mfg + '_Primary.jpg'
            urlretrieve(url, output_directory + f"\{primary_file_name}")
            file_name_dict['Primary_File_name'].append(primary_file_name)

        except HTTPError as err:
            if err.code == 404:
                file_name_dict['404_error_images'].append(mfg)
                pass
            else:
                raise
        
        # except:
        #     file_name_dict['Primary_File_name'].append('NULL')
        #     file_name_dict['Sku'].append(mfg)

    # for mfg, url in img_2_dict.items():
    #     try:
    #         file_name2 = mfg + '_img2.jpg'
    #         urlretrieve(url, output_directory + f"\{file_name2}")
    #         file_name_dict['Image_2_Name'].append(file_name2)

    #     except HTTPError as err:
    #         if err.code == 404:
    #             file_name_dict['404_error_images'].append(mfg)
    #             pass
    #         else:
    #             raise
        
        # except:
        #     file_name_dict['Image_2_Name'].append('NULL')
        #     file_name_dict['Sku'].append(mfg)

    # for mfg, url in img_3_dict.items():
    #     try:
    #         file_name3 = mfg + '_img3.jpg'
    #         urlretrieve(url, output_directory + f"\{file_name3}")
    #         file_name_dict['Image_3_Name'].append(file_name3)

    #     except HTTPError as err:
    #         if err.code == 404:
    #             file_name_dict['404_error_images'].append(mfg)
    #             pass
    #         else:
    #             raise
        
        # except:
        #     file_name_dict['Image_3_Name'].append('NULL')
        #     file_name_dict['Sku'].append(mfg)

    # for mfg, url in img_4_dict.items():
    #     try:
    #         file_name4 = mfg + '_img4.jpg'
    #         urlretrieve(url, output_directory + f"\{file_name4}")
    #         file_name_dict['Image_4_Name'].append(file_name4)

    #     except HTTPError as err:
    #         if err.code == 404:
    #             file_name_dict['404_error_images'].append(mfg)
    #             pass
    #         else:
    #             raise
        
        # except:
        #     file_name_dict['Image_4_Name'].append('NULL')
        #     file_name_dict['Sku'].append(mfg)

    # Creating a dataframe from the file name list
    file_df = pd.DataFrame.from_dict(file_name_dict, orient='index')
    file_df = file_df.transpose()
    

    #  Writing the dataframe to an excel worksheet
    # path = r'C:\Users\ABG2137\OneDrive - MyDigitalSpace\Desktop\File_Data.xlsx'
    # excel_wb = load_workbook(path)
    file_df.to_excel('File_Data.xlsx', sheet_name='File_Data')

    # sg.popup("Run Complete!")

def make_main_window():
    
    # Theme of windows
    sg.theme('Dark Grey 13')
    
    # Creating window layouts
    main_layout = [[sg.Text("Team Product Tool")], 
                    [sg.Text("Tools will still run if IDs are mixed. Please ensure all IDs are in the right place.", text_color='red', font=('Arial Bold', 10))],
                    [sg.Text("Choose which tool you want.")],
                    [sg.Button("Build Tool"), sg.Button("Ferg Tool"), sg.Button("Image Converter Tool"), sg.Button("Exit")]]

    return sg.Window('Main Window', main_layout)

def make_build_window():
     # Theme of windows
    sg.theme('Dark Grey 13')

    img_layout = [[sg.Text("Build Scraper Tool")],
                  [sg.Text("Tools will still run if IDs are mixed. Please ensure all IDs are in the right place.", text_color='red', font=('Arial Bold', 10))],
                  [sg.Text('Please enter Sku(MFG Number) list.'), sg.InputText(key='-SKU-', pad=(0,0))],
                  [sg.Text('Please enter first ID list.'), sg.InputText(key='-ID_1-', pad=(0,0))],
                  [sg.Text('Please enter second ID list.'), sg.InputText(key='-ID_2-', pad=(0,0))],
                #   [sg.Text('Please enter the absolute path of Excel file to use.'), sg.InputText(key='-E_NAME-')],
                  [sg.Button("Run"), sg.Button("Exit")]]

    image_window = sg.Window('Build Scraper Window', img_layout, modal=True)

    while True:
        
        event, values = image_window.read()
        
        if event in(sg.WIN_CLOSED, "Exit"):
            break
        
        sku_list = values['-SKU-'].split('\n')
        id_1_list = values['-ID_1-'].split('\n')
        id_2_list = values['-ID_2-'].split('\n')
        # file_name = values['-E_NAME-'].rstrip()
        
        if event == 'Run':
            
            try:
                Build_tool(sku_list, id_1_list, id_2_list)
                sg.popup("Run Complete!")
            except:
                sg.popup("Something went wrong. Please make sure everything was entered correctly.")

    image_window.close()

def make_ferg_window():
     # Theme of windows
    sg.theme('Dark Grey 13')

    img_layout = [[sg.Text("Ferg Scraper Tool")],
                  [sg.Text('Please enter Alt sku list.'), sg.InputText(key='-ALTSKU-', pad=(0,0))],
                  [sg.Button("Run"), sg.Button("Exit")]]

    image_window = sg.Window('Ferg Scraper Window', img_layout, modal=True)

    while True:
        
        event, values = image_window.read()
        
        if event in(sg.WIN_CLOSED, "Exit"):
            break
        
        alt_sku_list = values['-ALTSKU-'].split('\n')
        
        if event == 'Run':
            
            try:
                Ferg_Tool(alt_sku_list)
                sg.popup("Run Complete!")
            except:
                sg.popup("Something went wrong. Please make sure everything was entered correctly.")

    image_window.close()

def make_converter_window():
     # Theme of windows
    sg.theme('Dark Grey 13')
    # WORK HERE
    converter_layout = [[sg.Text("Image Converter Tool")],
                        [sg.Text("Enter Skus and URLs for Primary images:")],
                        [sg.Text('Please enter Primary Sku(MFG Number) list.'), sg.InputText(key='-SKU-', pad=(0,0))],
                        [sg.Text('Please enter Primary image URL list.'), sg.InputText(key='-URL-', pad=(0,0))],
                        # [sg.Text("Enter Skus and URLs for Second image:")],
                        # [sg.Text('Please enter Sku(MFG Number) list for second image.'), sg.InputText(key='-SKU2-', pad=(0,0))],
                        # [sg.Text('Please enter image URL list for second image.'), sg.InputText(key='-URL2-', pad=(0,0))],
                        # [sg.Text("Enter Skus and URLs for Third image:")],
                        # [sg.Text('Please enter Sku(MFG Number) list for third image.'), sg.InputText(key='-SKU3-', pad=(0,0))],
                        # [sg.Text('Please enter image URL list for third image.'), sg.InputText(key='-URL3-', pad=(0,0))],
                        # [sg.Text("Enter Skus and URLs for Fourth image:")],
                        # [sg.Text('Please enter Sku(MFG Number) list for fourth image.'), sg.InputText(key='-SKU4-', pad=(0,0))],
                        # [sg.Text('Please enter image URL list for fourth image.'), sg.InputText(key='-URL4-', pad=(0,0))],
                        # [sg.Text('Please enter the absolute path of Excel file to use.'), sg.InputText(key='-E_NAME-')],
                        [sg.Text('Please enter the absolute path of folder to download images to.'), sg.InputText(key='-F_NAME-')],
                        [sg.Button("Run"), sg.Button("Exit")]]
    
    convert_window = sg.Window('Converter Window', converter_layout, modal=True)
    
    while True:
        
        event, values = convert_window.read()
        
        if event in(sg.WIN_CLOSED, "Exit"):
            break
        
        mfg_list_primary = values['-SKU-'].split('\n')
        # mfg_list_2 = values['-SKU2-'].split('\n')
        # mfg_list_3 = values['-SKU3-'].split('\n')
        # mfg_list_4 = values['-SKU4-'].split('\n')
        Primary_list = values['-URL-'].split('\n')
        # img_2_list = values['-URL2-'].split('\n')
        # img_3_list = values['-URL3-'].split('\n')
        # img_4_list = values['-URL4-'].split('\n')
        # excel_file_name = r'{}'.format(values['-E_NAME-'].rstrip())
        folder_name = r'{}'.format(values['-F_NAME-'].rstrip())

        if event == 'Run':
            
            try:
                # converter_tool(mfg_list_primary, mfg_list_2, mfg_list_3, mfg_list_4, Primary_list, img_2_list, img_3_list, img_4_list, folder_name)
                converter_tool(mfg_list_primary, Primary_list, folder_name)
                sg.popup("Run Complete!")
            except:
                sg.popup("Something went wrong. Please make sure everything was entered correctly.")

    convert_window.close()

def main():
    # Theme of windows
    sg.theme('Dark Grey 13')
    
    # Creating window layouts
    main_layout = [[sg.Text("Build Scraper Tool")], 
                    [sg.Text("Tools will still run if IDs are mixed. Please ensure all IDs are in the right place.", text_color='red', font=('Arial Bold', 10))],
                    [sg.Text("Choose which tool you want.")],
                    [sg.Button("Build Tool"), sg.Button("Ferg Tool"), sg.Button("Image Converter Tool"), sg.Button("Exit")]]

    main_window = sg.Window('Main Window', main_layout)

    # Event Loop
    while True:
        event, values = main_window.read()

        
        # End program if conditions met
        if event in(sg.WIN_CLOSED, "Exit"):
                break
        
        # Runs the Image scraper tool window and tool
        elif event == 'Build Tool':
            make_build_window()
        
        # Runs the Bullet scraper tool window and tool
        elif event == 'Ferg Tool':
            make_ferg_window()

        # Runs the Image URL Converter tool window and tool
        elif event == 'Image Converter Tool':
            make_converter_window()
    
    main_window.close()

# Run the program
if __name__ == "__main__":
    main()